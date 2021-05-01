# SPDX-License-Identifier: LGPL-3.0-only

"""Functions to export documents and items."""

import datetime
import os
import pathlib
from collections import defaultdict
from typing import Any, Dict

import openpyxl
import yaml

from doorstop import common, settings
from doorstop.common import DoorstopError
from doorstop.core.types import iter_documents, iter_items
from doorstop.core.tree import Tree
import portableqda,html

LIST_SEP = '\n'  # string separating list values when joined in a string

XLSX_MAX_WIDTH = 65.0  # maximum width for a column
XLSX_FILTER_PADDING = 3.5  # column padding to account for filter button

log = common.logger(__name__)

portableqda.log=log

try:
    CATEGORY_SEP = str(portableqda.CATEGORY_SEP)
except Exception:
    CATEGORY_SEP = "::"

codebook = None #coodebook to work on when exporting to QDC... TODO: is a global really needed??



def export(obj, path, ext=None, **kwargs):
    """Export an object to a given format.

    The function can be called in two ways:

    1. document or item-like object + output file path
    2. tree-like object + output directory path

    :param obj: (1) Item, list of Items, Document or (2) Tree
    :param path: (1) output file path or (2) output directory path
    :param ext: file extension to override output extension

    :raises: :class:`doorstop.common.DoorstopError` for unknown file formats

    :return: output location if files created, else None

    """
    # Determine the output format
    ext = ext or os.path.splitext(path)[-1] or '.csv'
    check(ext)
    if "whole_tree" in kwargs.keys() and kwargs["whole_tree"]:
        # Export the whole tree to one document
        kwargs["total_documents"] = len(obj.documents)
        log.info("exporting whole tree to {}...".format(path))
        count = 0
        pathNonlocal=(path+".")[:-1] #nonlocal-ish references the onefile object until closed
        if id(pathNonlocal) == id(path):
            raise ValueError("pathNonlocal should be a copy of path. Please report this to the develpers,"
                             "with as much information on the interpreter as you can")
        for obj2, path2 in iter_documents(obj, path, ext):
            count += 1
            kwargs["count"]=count
            # Export content to the specified path
            # common.create_dirname(path2)
            pathNonlocal = export_file(obj2, pathNonlocal, ext, **kwargs)

    else:
        # Export individual documents
        kwargs["total_documents"] = 1
        kwargs["count"] = 1
        check(ext)
        count = 0
        for obj2, path2 in iter_documents(obj, path, ext):
            count += 1

            # Export content to the specified path
            common.create_dirname(path2)
            log.info("exporting to {}...".format(path2))
            if ext in FORMAT_LINES:
                lines = export_lines(obj2, ext, **kwargs)
                common.write_lines(lines, path2)
            else:
                export_file(obj2, path2, ext, **kwargs)

    # Return the exported path
    if count:
        msg = "exported to {} file{}".format(count, 's' if count > 1 else '')
        log.info(msg)
        return path
    else:
        log.warning("nothing to export")
        return None


def export_lines(obj, ext='.yml', **kwargs):
    """Yield lines for an export in the specified format.

    :param obj: Item, list of Items, or Document to export
    :param ext: file extension to specify the output format

    :raises: :class:`doorstop.common.DoorstopError` for unknown file formats

    :return: lines generator

    """
    gen = check(ext, get_lines_gen=True)
    log.debug("yielding {} as lines of {}...".format(obj, ext))
    yield from gen(obj, **kwargs)


def export_file(obj, path, ext=None, **kwargs):
    """Create a file object for an export in the specified format.
    if "all" argument entered, check() will try to produce function that creates one file, otherwise each doc
    will have its own file.

    :param obj: Item, list of Items, or Document to export
    :param path: output file location with desired extension
    :param ext: file extension to override output path's extension

    :raises: :class:`doorstop.common.DoorstopError` for unknown file formats

    :return: path to created file

    """
    ext = ext or os.path.splitext(path)[-1]
    func = check(ext, get_file_func=True, **kwargs)
    log.debug("converting %s to file format %s...", obj, ext)
    try:
        #support for whole-project exporters (QDC, and some day QDPX and ReqIF)
        if 'whole_tree' in kwargs.keys():
            if not kwargs["whole_tree"]:
                del kwargs["whole_tree"]
                if 'total_documents' in kwargs.keys():
                    del kwargs["total_documents"]
                if 'count' in kwargs.keys():
                    del kwargs["count"]
        else:
            if 'total_documents' in kwargs.keys():
                del kwargs["total_documents"]
            if 'count' in kwargs.keys():
                del kwargs["count"]

        return func(obj, path, **kwargs)

    except IOError:
        msg = "unable to write to: {}".format(path)
        raise common.DoorstopFileError(msg) from None


def _lines_yaml(obj, **_):
    """Yield lines for a YAML export.

    :param obj: Item, list of Items, or Document to export

    :return: iterator of lines of text

    """
    for item in iter_items(obj):

        data = {str(item.uid): item.data}
        text = yaml.dump(data, default_flow_style=False, allow_unicode=True)
        yield text


def _tabulate(obj, sep=LIST_SEP, auto=False):
    """Yield lines of header/data for tabular export.

    :param obj: Item, list of Items, or Document to export
    :param sep: string separating list values when joined in a string
    :param auto: include placeholders for new items on import

    :return: iterator of rows of data

    """

    header = ['level', 'text', 'ref', 'links']

    # 'at_least_one_ref' detects if at least one of the items still have a deprecated 'ref' field.
    # If there is none, 'ref' header is excluded from the headers and is not exported.
    at_least_one_ref = False
    for item in iter_items(obj):
        data = item.data

        for value in sorted(data.keys()):
            if value not in header:
                header.append(value)

        ref_value = data.get('ref')
        if ref_value:
            at_least_one_ref = True

    try:
        reference_index = header.index('references')

        # Inserting 'references' header after the 'ref' header.
        header.insert(3, header.pop(reference_index))

        if not at_least_one_ref:
            header.remove('ref')
    except ValueError:
        pass

    yield ['uid'] + header

    for item in iter_items(obj):
        data = item.data

        # Yield row
        row = [item.uid]
        for key in header:
            value = data.get(key)
            if key == 'level':
                # some levels are floats for YAML presentation
                value = str(value)
            elif key == 'links':
                # separate identifiers with a delimiter
                value = sep.join(uid.string for uid in item.links)
            elif key == 'references':
                if value is None:
                    value = ''
                else:
                    ref_strings = []
                    for ref_item in value:
                        ref_type = ref_item['type']
                        ref_path = ref_item['path']

                        ref_string = "type:{},path:{}".format(ref_type, ref_path)

                        if 'keyword' in ref_item:
                            keyword = ref_item['keyword']
                            ref_string += ",keyword:{}".format(keyword)

                        ref_strings.append(ref_string)
                    value = '\n'.join(ref_string for ref_string in ref_strings)
            elif isinstance(value, str) and key not in ('reviewed',):
                # remove sentence boundaries and line wrapping
                value = item.get(key)
            elif value is None:
                value = ''
            row.append(value)
        yield row

    # Yield placeholders for new items
    if auto:
        for _ in range(settings.PLACEHOLDER_COUNT):
            yield [settings.PLACEHOLDER]


def _file_csv(obj, path, delimiter=',', auto=False):
    """Create a CSV file at the given path.

    :param obj: Item, list of Items, or Document to export
    :param path: location to export CSV file
    :param delimiter: character to delimit fields
    :param auto: include placeholders for new items on import

    :return: path of created file

    """
    return common.write_csv(
        _tabulate(obj, auto=auto),
        path,
        delimiter=delimiter,
        newline='',
        encoding='utf-8',
    )


def _file_tsv(obj, path, auto=False):
    """Create a TSV file at the given path.

    :param obj: Item, list of Items, or Document to export
    :param path: location to export TSV file
    :param auto: include placeholders for new items on import

    :return: path of created file

    """
    return _file_csv(obj, path, delimiter='\t', auto=auto)


def _file_xlsx(obj, path, auto=False):
    """Create an XLSX file at the given path.

    :param obj: Item, list of Items, or Document to export
    :param path: location to export XLSX file
    :param auto: include placeholders for new items on import

    :return: path of created file

    """
    workbook = _get_xlsx(obj, auto)
    workbook.save(path)

    return path


def _get_xlsx(obj, auto):
    """Create an XLSX workbook object.

    :param obj: Item, list of Items, or Document to export
    :param auto: include placeholders for new items on import

    :return: new workbook

    """
    col_widths: Dict[Any, float] = defaultdict(float)
    col = 'A'

    # Create a new workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Populate cells
    for row, data in enumerate(_tabulate(obj, auto=auto), start=1):
        for col_idx, value in enumerate(data, start=1):
            cell = worksheet.cell(column=col_idx, row=row)

            # wrap text in every cell
            alignment = openpyxl.styles.Alignment(
                vertical='top', horizontal='left', wrap_text=True
            )
            cell.alignment = alignment
            # and bold header rows
            if row == 1:
                cell.font = openpyxl.styles.Font(bold=True)

            # convert incompatible Excel types:
            # http://pythonhosted.org/openpyxl/api.html#openpyxl.cell.Cell.value
            if isinstance(value, (int, float, datetime.datetime)):
                cell.value = value
            else:
                cell.value = str(value)

            # track cell width
            col_widths[col_idx] = max(col_widths[col_idx], _width(str(value)))

    # Add filter up to the last column
    col_letter = openpyxl.utils.get_column_letter(len(col_widths))
    worksheet.auto_filter.ref = "A1:%s1" % col_letter

    # Set column width based on column contents
    for col in col_widths:
        if col_widths[col] > XLSX_MAX_WIDTH:
            width = XLSX_MAX_WIDTH
        else:
            width = col_widths[col] + XLSX_FILTER_PADDING
        col_letter = openpyxl.utils.get_column_letter(col)
        worksheet.column_dimensions[col_letter].width = width

    # Freeze top row
    worksheet.freeze_panes = worksheet.cell(row=2, column=1)

    return workbook

def _file_qdc(obj, path, auto=False, **kwargs):
    """Create a QDC codebook (REFI-QDA as per http://qdasoftware.org/) file at the given path.

    :param obj: Item, list of Items, or Document to export
    :param path: location to export QDPX file
    :param auto: include placeholders for new items on import

    :return: path of created file

    """
    #workbook = _get_qdpx(obj, auto)
    #workbook.save(path)
    #from xml.sax.saxutils import XMLGenerator
    with open(path,mode="w") as fh:
        qdcOutput = portableqda.codebook(output=fh)
        for item in obj.items:
            guid="id{}".format(item.uid) #TODO: retreive guid if exists
            error,errorDesc,code=qdcOutput.codeOp(name="{}:{}".format(obj.DEFAULT_PREFIX,item.uid),
                                    guid=guid,
                                    op="create")
            code.description=item.text

        # http://schema.qdasoftware.org/versions/Codebook/v1.0/Codebook.xsd
        # output = XMLGenerator(fh, encoding='utf-8',
        #                       short_empty_elements=True)
        qdcOutput.writeQdc()

    return path

def _onefile_qdc(obj, path, auto=False, **kwargs):
    """Create one QDC codebook for the whole tree (REFI-QDA as per http://qdasoftware.org/) file at the given path.

    :param obj: Item, list of Items, or Document to export
    :param path: location to export QDC file
    :param auto: include placeholders for new items on import

    :return: path of created file

    """
    #workbook = _get_qdpx(obj, auto)
    #workbook.save(path)
    #from xml.sax.saxutils import XMLGenerator
    global codebook
    project_name = str(obj.tree.document.prefix)
    if kwargs["count"]==1:
        result = path  # path to output file on the first iteration, the whole object aftewards
        #first doc in, create onefilOutput object #"guid" in item.extended #item.relpath
        if not str(path).lower().endswith(".qdc"):
            path+=".qdc"
        #fh = open(path,mode="w")
        codebook = portableqda.codebookCls(output=path)
        description="project {}\n\nDocuments: \n{}".format(project_name,obj.tree.draw())
    else:
        description = "document {}{}{}\n\n other documents: \n {}".format(obj.tree.document.prefix,CATEGORY_SEP,obj,obj.tree.draw())#)obj.tree)
        result = codebook
    # create the codebook SetCls representing the current doc
    if True:
        GroupName = project_name + CATEGORY_SEP
    else:
        GroupName = ""
    GroupName += CATEGORY_SEP.join(pathlib.Path(obj.relpath[2:]).parts)# @/A/B/C -> A::B::C, platform-indept

    error,errorDesc,setQda=codebook.createElement(elementCls=portableqda.setCls, #codebook elements are codeCls or setClas
                                                name=GroupName,
                                                description=description.replace("<-",CATEGORY_SEP))
                                                #description=html.escape(description))
    if error:
        log.warning("codebook.createElement({}): return error: {}".format(GroupName,errorDesc))
    else:
        log.debug("codebook: created Set {} for doc {}".format(GroupName,obj))
    for item in obj.items:
        if "guid" in item.extended:
            guid=item.data["guid"]
        else:
            guid=None #item.guid
        if "color" in item.extended:
            color=item.data["color"]
        else:
            color=None #item.color
        if len(item.header) >0:
            item_header=" - "+item.header
        else:
            item_header = ""
        codeName = "{}{}{}{}".format(project_name,CATEGORY_SEP, item.uid, item_header)
        itemGroupList=[GroupName,]
        #itemGroupName = CATEGORY_SEP.join(pathlib.Path(obj.relpath[2:]).parts) #@/A/B/C -> A::B::C, platform-indept
        #itemGroupList.append(itemGroupName)
        for itemGroupName in itemGroupList:
            log.debug("adding item {}, to group {}".format(codeName, itemGroupName))
        # error,errorDesc,itemGroup = result.codeSetOp(name=itemGroupName,do=portableqda.op.RETRIEVE,
        #                                              codeSetElm=portableqda.elm.SET)

        error,errorDesc,code=codebook.createElement(elementCls=portableqda.codeCls, name=codeName, guid=guid,
                                                    description=item.text, sets=itemGroupList)
        if error:
            log.warning("codebook.createElement({}): return error: {}".format(codeName,errorDesc))

        if guid != code.attrib["guid"]:
            log.warning("item {}: GUID changed from '{}' to '{}', saving item. check that Reviewed attribute is updated".
                  format(item.uid, guid, code.attrib["guid"]))
            item.set_attributes({"guid":code.attrib["guid"]})

        if color != code.attrib["color"]:
            log.warning("item {}: color changed from '{}' to '{}', saving item. check that Reviewed attribute is updated".
                  format(item.uid, color, code.attrib["color"]))
            item.set_attributes({"color":code.attrib["color"]})

    if kwargs["count"]==kwargs["total_documents"]:
        codebook.writeQdcFile()

    return result

def _get_qdc(obj, auto):
    """Create an QDC codebook file (REFI-QDA as per http://qdasoftware.org/)

    :param obj: Item, list of Items, or Document to export
    :param auto: include placeholders for new items on import

    :return: new workbook

    """
    col_widths: Dict[Any, float] = defaultdict(float)
    col = 'A'

    # Create a new workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Populate cells
    for row, data in enumerate(_tabulate(obj, auto=auto), start=1):
        for col_idx, value in enumerate(data, start=1):
            cell = worksheet.cell(column=col_idx, row=row)

            # wrap text in every cell
            alignment = openpyxl.styles.Alignment(
                vertical='top', horizontal='left', wrap_text=True
            )
            cell.alignment = alignment
            # and bold header rows
            if row == 1:
                cell.font = openpyxl.styles.Font(bold=True)

            # convert incompatible Excel types:
            # http://pythonhosted.org/openpyxl/api.html#openpyxl.cell.Cell.value
            if isinstance(value, (int, float, datetime.datetime)):
                cell.value = value
            else:
                cell.value = str(value)

            # track cell width
            col_widths[col_idx] = max(col_widths[col_idx], _width(str(value)))

    # Add filter up to the last column
    col_letter = openpyxl.utils.get_column_letter(len(col_widths))
    worksheet.auto_filter.ref = "A1:%s1" % col_letter

    # Set column width based on column contents
    for col in col_widths:
        if col_widths[col] > XLSX_MAX_WIDTH:
            width = XLSX_MAX_WIDTH
        else:
            width = col_widths[col] + XLSX_FILTER_PADDING
        col_letter = openpyxl.utils.get_column_letter(col)
        worksheet.column_dimensions[col_letter].width = width

    # Freeze top row
    worksheet.freeze_panes = worksheet.cell(row=2, column=1)

    return workbook

def _file_qdpx(obj, path, auto=False):
    """Create a QDPX project file (REFI-QDA as per http://qdasoftware.org/) file at the given path.

    :param obj: Item, list of Items, or Document to export
    :param path: location to export QDPX file
    :param auto: include placeholders for new items on import

    :return: path of created file

    """
    #workbook = _get_qdpx(obj, auto)
    #workbook.save(path)
    from xml.sax.saxutils import XMLGenerator
    import sys
    class Tag: pass
    tags = list()
    for item in obj.items:
        tag = Tag()
        tag.id="id{}".format(item.uid)
        #tag.path="path{}".format(idx)
        tag.path="{}:{}".format(obj.DEFAULT_PREFIX,item.uid)
        tag.description=item.text
        tags.append(tag)


    # http://schema.qdasoftware.org/versions/Codebook/v1.0/Codebook.xsd
    with open(path,mode="w") as fh:
        output = XMLGenerator(fh, encoding='utf-8',
                              short_empty_elements=True)
        output.startDocument()
        output.startPrefixMapping(None, 'urn:QDA-XML:codebook:1.0')
        portableqda.write_codebook(tags, output)
        output.endPrefixMapping(None)
        output.endDocument()

    return path

def _get_qdpx(obj, auto):
    """Create an QDPX file (REFI-QDA as per http://qdasoftware.org/)

    :param obj: Item, list of Items, or Document to export
    :param auto: include placeholders for new items on import

    :return: new workbook

    """
    col_widths: Dict[Any, float] = defaultdict(float)
    col = 'A'

    # Create a new workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Populate cells
    for row, data in enumerate(_tabulate(obj, auto=auto), start=1):
        for col_idx, value in enumerate(data, start=1):
            cell = worksheet.cell(column=col_idx, row=row)

            # wrap text in every cell
            alignment = openpyxl.styles.Alignment(
                vertical='top', horizontal='left', wrap_text=True
            )
            cell.alignment = alignment
            # and bold header rows
            if row == 1:
                cell.font = openpyxl.styles.Font(bold=True)

            # convert incompatible Excel types:
            # http://pythonhosted.org/openpyxl/api.html#openpyxl.cell.Cell.value
            if isinstance(value, (int, float, datetime.datetime)):
                cell.value = value
            else:
                cell.value = str(value)

            # track cell width
            col_widths[col_idx] = max(col_widths[col_idx], _width(str(value)))

    # Add filter up to the last column
    col_letter = openpyxl.utils.get_column_letter(len(col_widths))
    worksheet.auto_filter.ref = "A1:%s1" % col_letter

    # Set column width based on column contents
    for col in col_widths:
        if col_widths[col] > XLSX_MAX_WIDTH:
            width = XLSX_MAX_WIDTH
        else:
            width = col_widths[col] + XLSX_FILTER_PADDING
        col_letter = openpyxl.utils.get_column_letter(col)
        worksheet.column_dimensions[col_letter].width = width

    # Freeze top row
    worksheet.freeze_panes = worksheet.cell(row=2, column=1)

    return workbook

def _width(text):
    """Get the maximum length in a multiline string."""
    if text:
        return max(len(line) for line in text.splitlines())
    else:
        return 0


# Mapping from file extension to lines generator
FORMAT_LINES = {'.yml': _lines_yaml}
# Mapping from file extension to file generator
FORMAT_FILE = {'.csv': _file_csv, '.tsv': _file_tsv, '.xlsx': _file_xlsx,  '.qdc': _file_qdc ,  '.qdpx': _file_qdpx }
# Mapping from file extension to one file generator
FORMAT_ONEFILE = {'.qdc': _onefile_qdc ,  '.qdpx': _onefile_qdc} #TODO: px }
# Union of format dictionaries
FORMAT = dict(list(FORMAT_LINES.items()) + list(FORMAT_FILE.items()))  # type: ignore


def check(ext, get_lines_gen=False, get_file_func=False, **kwargs):
    """Confirm an extension is supported for export.

    :param get_lines_func: return a lines generator if available
    :param get_file_func: return a file creator if available

    :raises: :class:`doorstop.common.DoorstopError` for unknown formats

    :return: function requested if available

    """
    exts = ', '.join(ext for ext in FORMAT)
    lines_exts = ', '.join(ext for ext in FORMAT_LINES)
    file_exts = ', '.join(ext for ext in FORMAT_FILE)
    fmt = "unknown {{}} format: {} (options: {{}})".format(ext or None)

    if get_lines_gen:
        try:
            gen = FORMAT_LINES[ext]
        except KeyError:
            exc = DoorstopError(fmt.format("lines export", lines_exts))
            raise exc from None
        else:
            log.debug("found lines generator for: {}".format(ext))
            return gen

    if get_file_func:
        #if "whole_tree" in kwargs.keys(): #TODO: check this line in upstream
        if "whole_tree" in kwargs.keys() and kwargs["whole_tree"]:
            try:
                func = FORMAT_ONEFILE[ext]
            except KeyError:
                exc = DoorstopError(fmt.format("one file export", file_exts))
                raise exc from None
            else:
                log.debug("found one file creator for: {}".format(ext))
                return func
        else:
            try:
                func = FORMAT_FILE[ext]
            except KeyError:
                exc = DoorstopError(fmt.format("file export", file_exts))
                raise exc from None
            else:
                log.debug("found file creator for: {}".format(ext))
                return func

    if ext not in FORMAT:
        exc = DoorstopError(fmt.format("export", exts))
        raise exc

    return None
